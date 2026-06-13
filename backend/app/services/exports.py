"""ExportService: generate XLSX, CSV, or PDF from entity records."""
from __future__ import annotations

import csv
import io
import uuid
from typing import Any

import structlog
from sqlalchemy.ext.asyncio import AsyncSession

logger = structlog.get_logger(__name__)

_MAX_EXPORT_ROWS = 10_000


class ExportError(Exception):
    def __init__(self, detail: str, status_code: int = 400) -> None:
        self.detail = detail
        self.status_code = status_code
        super().__init__(detail)


class ExportService:
    def __init__(self, db: AsyncSession) -> None:
        self._db = db

    async def export(
        self,
        entity_id: uuid.UUID,
        params: Any,  # RecordListParams
        format: str = "xlsx",
    ) -> bytes:
        """Fetch all matching records and return serialized file bytes."""
        from app.services.records import RecordService
        from app.schemas.records import RecordListParams

        # Override limit to export cap; fetch up to _MAX_EXPORT_ROWS
        export_params = params.model_copy(update={"limit": min(params.limit, _MAX_EXPORT_ROWS)})
        page = await RecordService(self._db).list_records(entity_id, export_params)
        records = page.items

        if not records:
            payloads: list[dict[str, Any]] = []
        else:
            payloads = [r.payload for r in records]

        # Collect all unique field names (preserve insertion order)
        headers: list[str] = []
        seen: set[str] = set()
        for p in payloads:
            for k in p:
                if k not in seen:
                    headers.append(k)
                    seen.add(k)

        logger.info("export_started", entity_id=str(entity_id), format=format,
                    row_count=len(payloads))

        match format:
            case "csv":
                return self._to_csv(headers, payloads)
            case "xlsx":
                return self._to_xlsx(headers, payloads)
            case "pdf":
                return self._to_pdf(headers, payloads)
            case _:
                raise ExportError(f"Unsupported export format: {format!r}")

    # ------------------------------------------------------------------
    # Serializers
    # ------------------------------------------------------------------

    @staticmethod
    def _to_csv(headers: list[str], rows: list[dict[str, Any]]) -> bytes:
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore",
                                lineterminator="\r\n")
        writer.writeheader()
        writer.writerows(rows)
        return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility

    @staticmethod
    def _to_xlsx(headers: list[str], rows: list[dict[str, Any]]) -> bytes:
        try:
            import openpyxl  # noqa: PLC0415
            from openpyxl.styles import Font, PatternFill  # noqa: PLC0415
        except ImportError as exc:
            raise ExportError("openpyxl is required for XLSX export", status_code=501) from exc

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Records"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")

        # Header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[cell.column_letter].width = max(12, len(header) + 4)

        # Data rows
        for row_idx, payload in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                val = payload.get(header)
                if isinstance(val, (list, dict)):
                    val = str(val)
                ws.cell(row=row_idx, column=col_idx, value=val)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _to_pdf(headers: list[str], rows: list[dict[str, Any]]) -> bytes:
        try:
            from fpdf import FPDF  # noqa: PLC0415
        except ImportError as exc:
            raise ExportError("fpdf2 is required for PDF export", status_code=501) from exc

        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=10)
        pdf.add_page()
        pdf.set_font("Helvetica", style="B", size=9)

        # Column width: distribute evenly, max 40mm each
        page_w = pdf.w - pdf.l_margin - pdf.r_margin
        col_w = min(40.0, page_w / max(len(headers), 1))
        row_h = 6.0

        # Header row
        for h in headers:
            pdf.set_fill_color(31, 78, 121)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(col_w, row_h, str(h)[:20], border=1, fill=True)
        pdf.ln()

        pdf.set_font("Helvetica", size=8)
        pdf.set_text_color(0, 0, 0)
        for payload in rows:
            for h in headers:
                val = payload.get(h)
                text = "" if val is None else str(val)[:30]
                pdf.set_fill_color(240, 240, 240)
                pdf.cell(col_w, row_h, text, border=1)
            pdf.ln()

        return bytes(pdf.output())
